import os
import openpyxl
from tkinter import messagebox, filedialog
from pdf2image import convert_from_path
from image_processing import process_image
from PIL import Image

def classify(root, pdf_image, crop_coords, pdf_files, rect_id, replace_var, user_entry_var, debug_images, status_label, canvas, show_recognition_results):
    """Classify the cropped region of the PDF image and store debug images."""
    if pdf_image[0] is None:
        messagebox.showerror("Error", "No PDF image loaded.")
        return
    if crop_coords[0] is None:
        messagebox.showerror("Error", "Please draw a bounding box first.")
        return
    
    status_label.config(text="Processing... Please wait.")
    root.update()

    if rect_id[0] is not None:
        canvas.delete(rect_id[0])
        rect_id[0] = None

    recognition_results = []
    file_paths = pdf_files
    
    for file_path in file_paths:
        try:
            images = convert_from_path(file_path, dpi=200)
            if len(images) != 1:
                recognition_results.append(f"{os.path.basename(file_path)}: Not a one-page PDF.")
                continue
            original_img = images[0]
            cropped_image = original_img.crop(crop_coords[0])
            temp_image_path = "temp_cropped_region.png"
            cropped_image.save(temp_image_path)
            
            debug_entry = {
                'original': original_img,
                'cropped': cropped_image,
                'processed': None,
                'mnist_images': None,
                'green_boxes_img': None,
                'binary_results': None
            }
            
            digit_results, processed_img, mnist_images, green_boxes_img, binary_results = process_image(temp_image_path)
            processed_img_pil = Image.fromarray(processed_img)
            green_boxes_img_pil = Image.fromarray(green_boxes_img)

            debug_entry.update({
                'processed': processed_img_pil,
                'mnist_images': mnist_images,
                'results': ''.join(digit_results),
                'green_boxes_img': green_boxes_img_pil,
                'binary_results': binary_results
            })
            
            recognition_results.append(f"{os.path.basename(file_path)}: {''.join(digit_results)}")
            
            debug_images[file_path] = debug_entry
            
            if os.path.exists(temp_image_path):
                os.remove(temp_image_path)
        except Exception as e:
            recognition_results.append(f"{os.path.basename(file_path)}: Error processing file - {e}")

    if replace_var.get() and user_entry_var.get().strip():
        user_input = user_entry_var.get().strip()
        modified_results = []
        for result in recognition_results:
            parts = result.split(': ', 1)
            if len(parts) == 2:
                filename, content = parts
                if content.isdigit():
                    content_list = list(content)
                    replace_length = min(len(user_input), len(content_list))
                    for i in range(replace_length):
                        content_list[i] = user_input[i]
                    new_content = ''.join(content_list)
                    modified_results.append(f"{filename}: {new_content}")
                else:
                    modified_results.append(result)
            else:
                modified_results.append(result)
        recognition_results = modified_results

        for file_path in file_paths:
            if file_path in debug_images:
                file_index = file_paths.index(file_path)
                if 0 <= file_index < len(recognition_results):
                    debug_images[file_path]['results'] = recognition_results[file_index].split(': ')[1]

    status_label.config(text="Classification complete!")
    show_recognition_results(recognition_results)

def export_to_excel(results, parent_window):
    """Export recognition results to an Excel file."""
    save_dir = filedialog.askdirectory(title="Select Folder to Save Excel File", parent=parent_window)
    if not save_dir:
        return
    
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Recognition Results"
        
        ws['A1'] = "File Name"
        ws['B1'] = "Recognized Text"
        
        for row, result in enumerate(results, start=2):
            parts = result.split(': ', 1)
            if len(parts) == 2:
                filename, text = parts
                ws[f'A{row}'] = filename
                ws[f'B{row}'] = text
            else:
                ws[f'A{row}'] = "Error"
                ws[f'B{row}'] = result
        
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        filename = os.path.join(save_dir, f"recognition_results_{timestamp}.xlsx")
        wb.save(filename)
        messagebox.showinfo("Export Complete", f"Results exported successfully to:\n{filename}")
    except Exception as e:
        messagebox.showerror("Export Error", f"Failed to export to Excel: {e}")