import os
import time
import openpyxl
from tkinter import filedialog, messagebox
from PIL import Image

def export_to_excel(results, parent_window):
    """Export recognition results to an Excel file.

    Args:
        results: List of recognition results as strings.
        parent_window: Parent Tkinter window for the file dialog.
    """
    save_dir = filedialog.askdirectory(title="Select Folder to Save Excel File", parent=parent_window)
    if not save_dir:
        return
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Recognition Results"
        ws['A1'] = "File Name"
        ws['B1'] = "Recognized Text"
        for row, result in enumerate(results, start=2):
            parts = result.split(': ', 1)
            if len(parts) == 2:
                filename, text = parts
                ws[f'A{row}'] = filename
                ws[f'B{row}'] = text
            else:
                ws[f'A{row}'] = "Error"
                ws[f'B{row}'] = result
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        filename = os.path.join(save_dir, f"recognition_results_{timestamp}.xlsx")
        wb.save(filename)
        messagebox.showinfo("Export Complete", f"Results exported successfully to:\n{filename}")
    except Exception as e:
        messagebox.showerror("Export Error", f"Failed to export to Excel: {e}")

def download_image(image, filename_prefix, parent_window, file_path=None):
    """Prompt user to save a single image with a unique filename.

    Args:
        image: PIL Image to save.
        filename_prefix: Prefix for the filename.
        parent_window: Parent Tkinter window for the file dialog.
        file_path: Optional path to the original file for naming.
    """
    if not image:
        messagebox.showinfo("Download Error", f"No {filename_prefix} image available to download.")
        return
    save_dir = filedialog.askdirectory(title=f"Select Folder to Save {filename_prefix} Image", parent=parent_window)
    if not save_dir:
        return
    try:
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        pdf_basename = os.path.splitext(os.path.basename(file_path))[0] if file_path else "unknown"
        filename = os.path.join(save_dir, f"{filename_prefix}_{pdf_basename}_{timestamp}.png")
        image.save(filename)
        messagebox.showinfo("Download Complete", f"{filename_prefix} image saved successfully to:\n{save_dir}")
    except Exception as e:
        messagebox.showerror("Download Error", f"Failed to save {filename_prefix} image: {e}")

def download_mnist_images(debug_data, parent_window, file_path=None):
    """Prompt user to save MNIST-formatted images with unique filenames.

    Args:
        debug_data: Dictionary containing MNIST images.
        parent_window: Parent Tkinter window for the file dialog.
        file_path: Optional path to the original file for naming.
    """
    if 'mnist_images' not in debug_data or not debug_data['mnist_images']:
        messagebox.showinfo("Download Error", "No MNIST-formatted images available to download.")
        return
    save_dir = filedialog.askdirectory(title="Select Folder to Save MNIST Images", parent=parent_window)
    if not save_dir:
        return
    try:
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        pdf_basename = os.path.splitext(os.path.basename(file_path))[0] if file_path else "unknown"
        for i, mnist_img in enumerate(debug_data['mnist_images']):
            mnist_img_pil = Image.fromarray(mnist_img)
            filename = os.path.join(save_dir, f"mnist_char_{pdf_basename}_{timestamp}_{i+1}.png")
            mnist_img_pil.save(filename)
        messagebox.showinfo("Download Complete", f"MNIST images saved successfully to:\n{save_dir}")
    except Exception as e:
        messagebox.showerror("Download Error", f"Failed to save MNIST images: {e}")