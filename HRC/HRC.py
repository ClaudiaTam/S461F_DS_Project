import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk
from PIL import Image, ImageTk
from pdf2image import convert_from_path
import torch
from torchvision import transforms
import openpyxl
import cv2
import numpy as np
import shutil
import time
import os

# Define the digit classification model architecture (10 classes: 0-9)
class DigitNet(torch.nn.Module):
    def __init__(self):
        super(DigitNet, self).__init__()
        self.conv1 = torch.nn.Conv2d(1, 64, 3, 1)
        self.conv2 = torch.nn.Conv2d(64, 128, 3, 1)
        self.conv3 = torch.nn.Conv2d(128, 256, 3, 1)
        self.dropout1 = torch.nn.Dropout(0.25)
        self.dropout2 = torch.nn.Dropout(0.4)
        self.fc1 = torch.nn.Linear(256 * 5 * 5, 512)
        self.fc2 = torch.nn.Linear(512, 256)
        self.fc3 = torch.nn.Linear(256, 10)

    def forward(self, x):
        x = self.conv1(x)
        x = torch.nn.functional.relu(x)
        x = self.conv2(x)
        x = torch.nn.functional.relu(x)
        x = torch.nn.functional.max_pool2d(x, 2)
        x = self.conv3(x)
        x = torch.nn.functional.relu(x)
        x = torch.nn.functional.max_pool2d(x, 2)
        x = self.dropout1(x)
        x = torch.flatten(x, 1)
        x = self.fc1(x)
        x = torch.nn.functional.relu(x)
        x = self.dropout2(x)
        x = self.fc2(x)
        x = torch.nn.functional.relu(x)
        x = self.fc3(x)
        return x

# Define the letter classification model architecture (26 classes: A-Z)
class LetterNet(torch.nn.Module):
    def __init__(self):
        super(LetterNet, self).__init__()
        self.conv1 = torch.nn.Conv2d(1, 64, 3, 1)
        self.conv2 = torch.nn.Conv2d(64, 128, 3, 1)
        self.conv3 = torch.nn.Conv2d(128, 256, 3, 1)
        self.dropout1 = torch.nn.Dropout(0.25)
        self.dropout2 = torch.nn.Dropout(0.4)
        self.fc1 = torch.nn.Linear(256 * 5 * 5, 512)
        self.fc2 = torch.nn.Linear(512, 256)
        self.fc3 = torch.nn.Linear(256, 26)  # 26 classes for letters A-Z

    def forward(self, x):
        x = self.conv1(x)
        x = torch.nn.functional.relu(x)
        x = self.conv2(x)
        x = torch.nn.functional.relu(x)
        x = torch.nn.functional.max_pool2d(x, 2)
        x = self.conv3(x)
        x = torch.nn.functional.relu(x)
        x = torch.nn.functional.max_pool2d(x, 2)
        x = self.dropout1(x)
        x = torch.flatten(x, 1)
        x = self.fc1(x)
        x = torch.nn.functional.relu(x)
        x = self.dropout2(x)
        x = self.fc2(x)
        x = torch.nn.functional.relu(x)
        x = self.fc3(x)
        return x

# Define the binary classification model architecture (2 classes: digit or letter)
class BinaryNet(torch.nn.Module):
    def __init__(self):
        super(BinaryNet, self).__init__()
        self.conv1 = torch.nn.Conv2d(1, 64, 3, 1)
        self.conv2 = torch.nn.Conv2d(64, 128, 3, 1)
        self.conv3 = torch.nn.Conv2d(128, 256, 3, 1)
        self.dropout1 = torch.nn.Dropout(0.25)
        self.dropout2 = torch.nn.Dropout(0.4)
        self.fc1 = torch.nn.Linear(256 * 5 * 5, 512)
        self.fc2 = torch.nn.Linear(512, 256)
        self.fc3 = torch.nn.Linear(256, 1)  # 2 outputs: digit (0) or letter (1)

    def forward(self, x):
        x = self.conv1(x)
        x = torch.nn.functional.relu(x)
        x = self.conv2(x)
        x = torch.nn.functional.relu(x)
        x = torch.nn.functional.max_pool2d(x, 2)
        x = self.conv3(x)
        x = torch.nn.functional.relu(x)
        x = torch.nn.functional.max_pool2d(x, 2)
        x = self.dropout1(x)
        x = torch.flatten(x, 1)
        x = self.fc1(x)
        x = torch.nn.functional.relu(x)
        x = self.dropout2(x)
        x = self.fc2(x)
        x = torch.nn.functional.relu(x)
        x = self.fc3(x)
        output = torch.sigmoid(x)
        return output

# Load the pre-trained digit classification model
digit_model = DigitNet()
digit_model_file = "/Users/ryan/Desktop/DL/FYP/Combined_Model_seed_53881/combined_cnn_epoch:14_test-accuracy:99.5865_test-loss:0.0144.pt"
try:
    digit_model.load_state_dict(torch.load(digit_model_file, map_location=torch.device('cpu')))
    digit_model.eval()
except Exception as e:
    print(f"Error loading digit model: {e}")
    exit()

# Load the pre-trained letter classification model
letter_model = LetterNet()
letter_model_file = "/Users/ryan/Desktop/DL/FYP/Capitals_Model_seed_76702/capitals_cnn_epoch_15_test-accuracy_97.7364_test-loss_0.0015.pt"  # Replace with the actual path to your trained letter model
try:
    letter_model.load_state_dict(torch.load(letter_model_file, map_location=torch.device('cpu')))
    letter_model.eval()
except Exception as e:
    print(f"Error loading letter model: {e}")
    exit()  

# Load the pre-trained binary classification model
binary_model = BinaryNet()
binary_model_file = "/Users/ryan/Desktop/DL/FYP/Binary_Model_seed_67675/binary_cnn_epoch:54_test-accuracy:99.9306_test-loss:0.0046.pt"  # Replace with your actual binary model path
try:
    binary_model.load_state_dict(torch.load(binary_model_file, map_location=torch.device('cpu')))
    binary_model.eval()
except Exception as e:
    print(f"Error loading binary model: {e}")
    exit()

# Image transformation for MNIST compatibility
transform = transforms.Compose([
    transforms.Grayscale(num_output_channels=1),
    transforms.Resize((28, 28)),
    transforms.ToTensor(),
    transforms.Normalize((0.1307,), (0.3081,))
])

def classify_image(image_path):
    """Classify an image with binary model first, then digit or letter model."""
    try:
        image = Image.open(image_path).convert("L")
        image_tensor = transform(image).unsqueeze(0)
        with torch.no_grad():
            # Binary classification first
            binary_output = binary_model(image_tensor)
            binary_pred = binary_output.argmax(dim=1, keepdim=True).item()
            binary_label = "Digit" if binary_pred == 0 else "Letter"

            # Route to appropriate model based on binary result
            if binary_pred == 0:  # Digit
                digit_output = digit_model(image_tensor)
                pred = digit_output.argmax(dim=1, keepdim=True).item()
                result = str(pred)
            else:  # Letter
                letter_output = letter_model(image_tensor)
                pred = letter_output.argmax(dim=1, keepdim=True).item()
                result = chr(65 + pred)  # Convert 0-25 to A-Z

        return result, binary_label
    except Exception as e:
        print(f"Error processing image: {e}")
        return None, None

def load_pdf():
    """Load and display multiple one-page PDF files."""
    global pdf_files, pdf_image, debug_images
    file_paths = filedialog.askopenfilenames(
        filetypes=[("PDF Files", "*.pdf")],
        title="Select one or more one-page PDF files"
    )
    if not file_paths:
        return
    pdf_files = list(file_paths)
    debug_images.clear()
    pdf_combobox['values'] = [os.path.basename(path) for path in pdf_files]
    if pdf_files:
        pdf_combobox.current(0)
        display_selected_pdf()

def display_selected_pdf():
    """Display the selected PDF from the combobox."""
    global pdf_image, pdf_files
    if not pdf_files:
        return
    
    selection = pdf_combobox.current()
    if selection < 0:
        return
        
    try:
        images = convert_from_path(pdf_files[selection], dpi=200)
        if len(images) != 1:
            messagebox.showerror("Error", "Please upload one-page PDF files.")
            return
        pdf_image = images[0]
        display_image(pdf_image)
        status_label.config(text=f"Loaded: {os.path.basename(pdf_files[selection])}")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to load PDF: {e}")

def display_image(image):
    """Display the given PIL image in the GUI with zoom support."""
    global canvas, tk_image, scaled_image, image_scale, canvas_offset_x, canvas_offset_y, zoom_factor, original_image
    canvas.delete("all")
    
    # Store the original image for zooming
    original_image = image
    
    image_width, image_height = image.size
    canvas_width, canvas_height = canvas.winfo_width(), canvas.winfo_height()
    if canvas_width <= 1:  # Canvas not yet drawn
        canvas_width, canvas_height = 600, 800
    
    # Calculate display scale based on both window size and zoom factor
    scale_factor = min(canvas_width / image_width, canvas_height / image_height) * zoom_factor
    image_scale = scale_factor
    
    new_width = int(image_width * scale_factor)
    new_height = int(image_height * scale_factor)
    scaled_image = image.resize((new_width, new_height))
    
    # Center the image on the canvas, considering scroll position
    canvas_offset_x = max(0, (canvas_width - new_width) // 2)
    canvas_offset_y = max(0, (canvas_height - new_height) // 2)
    
    tk_image = ImageTk.PhotoImage(scaled_image)
    canvas.create_image(canvas_offset_x, canvas_offset_y, anchor="nw", image=tk_image)
    
    # Configure the canvas scrollregion for panning when zoomed
    if zoom_factor > 1.0:
        canvas.configure(scrollregion=(0, 0, max(canvas_width, new_width), max(canvas_height, new_height)))
    else:
        canvas.configure(scrollregion=(0, 0, canvas_width, canvas_height))

def on_mousewheel(event):
    """Handle mouse wheel events for zooming."""
    global zoom_factor, original_image
    
    # Only zoom if in zoom mode
    if not zoom_mode or not original_image:
        return
    
    # Get the cursor position for zoom focus
    canvas_x = canvas.canvasx(event.x)
    canvas_y = canvas.canvasy(event.y)
    
    # Determine zoom direction
    if event.delta > 0 or (hasattr(event, 'num') and event.num == 4):
        # Zoom in (limit maximum zoom)
        zoom_factor = min(zoom_factor * 1.1, 5.0)
    else:
        # Zoom out (limit minimum zoom)
        zoom_factor = max(zoom_factor * 0.9, 0.5)
    
    # Redisplay with new zoom factor
    display_image(original_image)
    
    # Update status to show zoom level
    status_label.config(text=f"Zoom: {zoom_factor:.1f}x")

def start_pan(event):
    """Start panning the canvas when zoom level > 1."""
    global pan_start_x, pan_start_y
    
    if zoom_factor > 1.0:
        canvas.config(cursor="fleur")  # Change cursor to indicate panning
        pan_start_x = event.x
        pan_start_y = event.y

def update_pan(event):
    """Pan the canvas during mouse drag when zoomed in."""
    global pan_start_x, pan_start_y
    
    if zoom_factor > 1.0 and pan_start_x is not None:
        # Calculate the amount to move
        dx = pan_start_x - event.x
        dy = pan_start_y - event.y
        
        # Move the canvas view
        canvas.xview_scroll(dx, "units")
        canvas.yview_scroll(dy, "units")
        
        # Update the start position for the next move
        pan_start_x = event.x
        pan_start_y = event.y

def end_pan(event):
    """End panning and reset cursor."""
    global pan_start_x, pan_start_y
    
    canvas.config(cursor="arrow")
    pan_start_x = None
    pan_start_y = None

def toggle_zoom_mode():
    """Toggle between crop and zoom mode."""
    global zoom_mode
    zoom_mode = not zoom_mode
    
    if zoom_mode:
        # Remove crop bindings and add zoom/pan bindings
        canvas.unbind("<ButtonPress-1>")
        canvas.unbind("<B1-Motion>")
        canvas.unbind("<ButtonRelease-1>")
        
        # Add bindings for panning
        canvas.bind("<ButtonPress-1>", start_pan)
        canvas.bind("<B1-Motion>", update_pan)
        canvas.bind("<ButtonRelease-1>", end_pan)
        
        # Ensure mousewheel bindings are active for zoom mode
        canvas.bind("<MouseWheel>", on_mousewheel)  # For Windows and macOS
        canvas.bind("<Button-4>", on_mousewheel)    # For Linux scroll up
        canvas.bind("<Button-5>", on_mousewheel)    # For Linux scroll down
        
        zoom_button.config(text="Switch to Crop Mode", bg="#FF9800")
        status_label.config(text="Zoom Mode: Use trackpad/wheel to zoom, click and drag to pan")
    else:
        # Remove pan and zoom bindings
        canvas.unbind("<ButtonPress-1>")
        canvas.unbind("<B1-Motion>")
        canvas.unbind("<ButtonRelease-1>")
        canvas.unbind("<MouseWheel>")
        canvas.unbind("<Button-4>")
        canvas.unbind("<Button-5>")
        
        # Add crop bindings
        canvas.bind("<ButtonPress-1>", start_crop)
        canvas.bind("<B1-Motion>", update_crop)
        canvas.bind("<ButtonRelease-1>", finish_crop)
        
        zoom_button.config(text="Switch to Zoom Mode", bg="#2196F3")
        status_label.config(text="Crop Mode: Click and drag to select an area")

def reset_zoom():
    """Reset zoom to 100%."""
    global zoom_factor, original_image
    zoom_factor = 1.0
    if original_image:
        display_image(original_image)
    status_label.config(text="Zoom reset to 100%")

def start_crop(event):
    """Start selecting rectangle area for cropping."""
    global start_x, start_y, rect_id
    if rect_id is not None:
        canvas.delete(rect_id)
        rect_id = None
    
    # Get current canvas scroll position
    canvas_x = canvas.canvasx(event.x)
    canvas_y = canvas.canvasy(event.y)
    
    # Adjust for both offset and scroll position
    start_x = canvas_x - canvas_offset_x
    start_y = canvas_y - canvas_offset_y
    
    rect_id = canvas.create_rectangle(canvas_x, canvas_y, canvas_x, canvas_y, outline="red", width=2)

def update_crop(event):
    global rect_id
    canvas_x = canvas.canvasx(event.x)
    canvas_y = canvas.canvasy(event.y)
    
    # Use canvas coordinates for the rectangle display
    canvas.coords(rect_id, start_x + canvas_offset_x, start_y + canvas_offset_y, canvas_x, canvas_y)

def finish_crop(event):
    global crop_coords, rect_id
    if pdf_image is None:
        messagebox.showerror("Error", "No PDF image loaded.")
        return

    try:
        canvas_x = canvas.canvasx(event.x)
        canvas_y = canvas.canvasy(event.y)
        
        # Calculate end coordinates in image space
        end_x = canvas_x - canvas_offset_x
        end_y = canvas_y - canvas_offset_y
        
        # Find min/max for proper rectangle regardless of drag direction
        x1, y1, x2, y2 = min(start_x, end_x), min(start_y, end_y), max(start_x, end_x), max(start_y, end_y)
        
        # Ensure coordinates are within scaled image bounds
        x1 = max(0, min(x1, scaled_image.width))
        y1 = max(0, min(y1, scaled_image.height))
        x2 = max(0, min(x2, scaled_image.width))
        y2 = max(0, min(y2, scaled_image.height))
        
        # Convert back to original image coordinates
        x1_orig = int(x1 / image_scale)
        y1_orig = int(y1 / image_scale)
        x2_orig = int(x2 / image_scale)
        y2_orig = int(y2 / image_scale)
        
        crop_coords = (x1_orig, y1_orig, x2_orig, y2_orig)
        status_label.config(text=f"Selection complete. Region: {crop_coords}")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to calculate crop coordinates: {e}")

def process_image(image_path):
    """Process the image, detect characters, crop to MNIST format, and classify them."""
    img = cv2.imread(image_path)

     # Step 1: Convert to grayscale
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
    enhanced_gray = clahe.apply(gray)

    # Step 2: Adaptive thresholding
    binary = cv2.adaptiveThreshold(
        enhanced_gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY_INV, 21, 12
    )

    # Step 3: Remove vertical and horizontal lines
    vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 30))
    eroded = cv2.erode(binary, vertical_kernel, iterations=1)
    detected_vertical_lines = cv2.morphologyEx(eroded, cv2.MORPH_OPEN, vertical_kernel)
    dilated_vertical_lines = cv2.dilate(detected_vertical_lines, vertical_kernel, iterations=1)
    horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (30, 1))
    eroded_h = cv2.erode(binary, horizontal_kernel, iterations=1)
    detected_horizontal_lines = cv2.morphologyEx(eroded_h, cv2.MORPH_OPEN, horizontal_kernel)
    dilated_horizontal_lines = cv2.dilate(detected_horizontal_lines, horizontal_kernel, iterations=1)
    combined_lines = cv2.add(dilated_vertical_lines, dilated_horizontal_lines)
    cleaned_binary = cv2.subtract(binary, combined_lines)

     # Step 4: Clean up noise
    closing_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))
    cleaned_binary = cv2.morphologyEx(cleaned_binary, cv2.MORPH_CLOSE, closing_kernel)
    cleaned_binary = cv2.medianBlur(cleaned_binary, 3)

    def preprocess_to_mnist_format(image):
        """Convert an image to MNIST format (28x28, centered, grayscale)."""
        if len(image.shape) > 2:
            image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        _, image = cv2.threshold(image, 128, 255, cv2.THRESH_BINARY)
        coords = cv2.findNonZero(image)
        if coords is None:
            return np.zeros((28, 28), dtype=np.uint8)
        x, y, w, h = cv2.boundingRect(coords)
        if w == 0 or h == 0:
            return np.zeros((28, 28), dtype=np.uint8)
        digit = image[y:y+h, x:x+w]
        max_dim = max(w, h)
        scale = 20.0 / max_dim
        new_w = int(w * scale)
        new_h = int(h * scale)
        resized = cv2.resize(digit, (new_w, new_h), interpolation=cv2.INTER_LINEAR)
        canvas = np.zeros((28, 28), dtype=np.uint8)
        x_offset = (28 - new_w) // 2
        y_offset = (28 - new_h) // 2
        canvas[y_offset:y_offset+new_h, x_offset:x_offset+new_w] = resized
        return canvas

    margin = 5
    contours, _ = cv2.findContours(cleaned_binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    output_dir = "/Users/ryan/Desktop/cropped_digits"
    os.makedirs(output_dir, exist_ok=True)

    img_height, img_width = cleaned_binary.shape

    def merge_nearby_contours(contours, distance_threshold=10):
        bounding_boxes = [cv2.boundingRect(c) for c in contours]
        merged_boxes = []
        used = [False] * len(bounding_boxes)

        for i in range(len(bounding_boxes)):
            if used[i]:
                continue
            x1, y1, w1, h1 = bounding_boxes[i]
            merged_xmin = x1
            merged_ymin = y1
            merged_xmax = x1 + w1
            merged_ymax = y1 + h1
            used[i] = True

            for j in range(i + 1, len(bounding_boxes)):
                if used[j]:
                    continue
                x2, y2, w2, h2 = bounding_boxes[j]
                dx = min(abs(x1 + w1 - x2), abs(x2 + w2 - x1)) if x1 < x2 + w2 and x2 < x1 + w1 else abs(x1 - x2)
                dy = min(abs(y1 + h1 - y2), abs(y2 + h2 - y1)) if y1 < y2 + h2 and y2 < y1 + h1 else abs(y1 - y2)
                distance = max(dx, dy)
                
                if distance < distance_threshold:
                    merged_xmin = min(merged_xmin, x2)
                    merged_ymin = min(merged_ymin, y2)
                    merged_xmax = max(merged_xmax, x2 + w2)
                    merged_ymax = max(merged_ymax, y2 + h2)
                    used[j] = True

            merged_boxes.append((merged_xmin, merged_ymin, merged_xmax - merged_xmin, merged_ymax - merged_ymin))
        return merged_boxes

    initial_boxes = []
    for contour in contours:
        x, y, w, h = cv2.boundingRect(contour)
        if 2 < w < 100 and 15 < h < 100:
            if (x > margin and y > margin and 
                x + w < img_width - margin and 
                y + h < img_height - margin):
                initial_boxes.append(contour)

    bounding_boxes = merge_nearby_contours(initial_boxes, distance_threshold=10)
    bounding_boxes.sort(key=lambda box: box[0])

    digit_results = []
    binary_results = []
    mnist_images = []
    green_boxes_img = cv2.cvtColor(cleaned_binary.copy(), cv2.COLOR_GRAY2BGR)

    for i, (x, y, w, h) in enumerate(bounding_boxes):
        digit_roi = cleaned_binary[y:y+h, x:x+w]
        mnist_digit = preprocess_to_mnist_format(digit_roi)
        mnist_images.append(mnist_digit)
        digit_filename = os.path.join(output_dir, f"char_{i+1}_mnist.png")
        cv2.imwrite(digit_filename, mnist_digit)
        result, binary_label = classify_image(digit_filename)
        if result is not None:
            digit_results.append(result)
            binary_results.append(binary_label)
        cv2.rectangle(green_boxes_img, (x, y), (x + w, y + h), (0, 255, 0), 2)

    return digit_results, cleaned_binary, mnist_images, green_boxes_img, binary_results, bounding_boxes

def classify():
    """Classify the cropped region of the PDF image and store debug images."""
    global crop_coords, pdf_files, rect_id, debug_images
    if pdf_image is None:
        messagebox.showerror("Error", "No PDF image loaded.")
        return
    if crop_coords is None:
        messagebox.showerror("Error", "Please draw a bounding box first.")
        return
    
    status_label.config(text="Processing... Please wait.")
    root.update()

    if rect_id is not None:
        canvas.delete(rect_id)
        rect_id = None

    recognition_results = []
    file_paths = pdf_files
    output_dir = "/Users/ryan/Desktop/cropped_digits"  # Define the folder path here
    
    for file_path in file_paths:
        try:
            images = convert_from_path(file_path, dpi=200)
            if len(images) != 1:
                recognition_results.append(f"{os.path.basename(file_path)}: Not a one-page PDF.")
                continue
            original_img = images[0]
            cropped_image = original_img.crop(crop_coords)
            temp_image_path = "temp_cropped_region.png"
            cropped_image.save(temp_image_path)
            
            debug_entry = {
                'original': original_img,
                'cropped': cropped_image,
                'processed': None,
                'mnist_images': None,
                'green_boxes_img': None,
                'binary_results': None
            }
            
            digit_results, processed_img, mnist_images, green_boxes_img, binary_results, bounding_boxes = process_image(temp_image_path)
            processed_img_pil = Image.fromarray(processed_img)
            green_boxes_img_pil = Image.fromarray(green_boxes_img)

            debug_entry.update({
                'processed': processed_img_pil,
                'mnist_images': mnist_images,
                'results': ''.join(digit_results),
                'green_boxes_img': green_boxes_img_pil,
                'binary_results': binary_results,
                'bounding_boxes': bounding_boxes  # Store the bounding boxes
            })
            
            recognition_results.append(f"{os.path.basename(file_path)}: {''.join(digit_results)}")
            
            debug_images[file_path] = debug_entry
            
            if os.path.exists(temp_image_path):
                os.remove(temp_image_path)
        except Exception as e:
            recognition_results.append(f"{os.path.basename(file_path)}: Error processing file - {e}")

    # Delete the "cropped_digits" folder after all processing is done
    if os.path.exists(output_dir):
        shutil.rmtree(output_dir)

    status_label.config(text="Classification complete!")
    show_recognition_results(recognition_results)

def export_to_excel(results, parent_window):
    """Export recognition results to an Excel file."""
    save_dir = filedialog.askdirectory(title="Select Folder to Save Excel File", parent=parent_window)
    if not save_dir:
        return
    
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Recognition Results"
        
        # Add headers
        ws['A1'] = "File Name"
        ws['B1'] = "Recognized Text"
        
        # Add data
        for row, result in enumerate(results, start=2):
            parts = result.split(': ', 1)
            if len(parts) == 2:
                filename, text = parts
                ws[f'A{row}'] = filename
                ws[f'B{row}'] = text
            else:
                ws[f'A{row}'] = "Error"
                ws[f'B{row}'] = result
        
        # Save the file with a timestamp to avoid overwriting
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        filename = os.path.join(save_dir, f"recognition_results_{timestamp}.xlsx")
        wb.save(filename)
        messagebox.showinfo("Export Complete", f"Results exported successfully to:\n{filename}")
    except Exception as e:
        messagebox.showerror("Export Error", f"Failed to export to Excel: {e}")

def show_recognition_results(results):
    """Display recognition results in a scrollable text window with an output button."""
    global result_window
    
    # If result_window exists, destroy it to create a fresh one
    if result_window is not None and result_window.winfo_exists():
        result_window.destroy()
    
    result_window = tk.Toplevel(root)
    result_window.title("Recognition Results")
    result_window.geometry("1200x400")
    
    control_frame = tk.Frame(result_window)
    control_frame.pack(fill=tk.X, padx=10, pady=5)
    
    pdf_label = tk.Label(control_frame, text="Select PDF for Debug:")
    pdf_label.pack(side=tk.LEFT, padx=5)
    
    pdf_debug_combobox = ttk.Combobox(control_frame, width=40, state="readonly")
    pdf_debug_combobox.pack(side=tk.LEFT, padx=5)
    pdf_debug_combobox['values'] = [os.path.basename(path) for path in pdf_files]
    if pdf_files:
        pdf_debug_combobox.current(pdf_combobox.current())
    
    def on_debug_select():
        selected_index = pdf_debug_combobox.current()
        if selected_index >= 0:
            show_debug_window(pdf_files[selected_index])
    
    debug_btn = tk.Button(control_frame, text="Show Debug Details", command=on_debug_select)
    debug_btn.pack(side=tk.LEFT, padx=5)

    output_btn = tk.Button(
        control_frame, 
        text="Output to Excel", 
        command=lambda: export_to_excel(results, result_window)
    )
    output_btn.pack(side=tk.LEFT, padx=5)
    
    close_btn = tk.Button(control_frame, text="Close", command=result_window.destroy)
    close_btn.pack(side=tk.RIGHT, padx=5)
    
    text_area = scrolledtext.ScrolledText(result_window, wrap=tk.WORD, width=60, height=20, font=("Arial", 14))
    text_area.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
    
    for result in results:
        text_area.insert(tk.END, result + "\n")
    text_area.configure(state="disabled")
    
    # Store the text_area in the window for later updates
    result_window.text_area = text_area

# Add this new helper function above `show_debug_window`
def download_image(image, filename_prefix, parent_window, file_path=None):
    """Prompt user to save a single image with a unique filename."""
    if not image:
        messagebox.showinfo("Download Error", f"No {filename_prefix} image available to download.")
        return
    
    save_dir = filedialog.askdirectory(title=f"Select Folder to Save {filename_prefix} Image", parent=parent_window)
    if not save_dir:
        return
    
    try:
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        pdf_basename = os.path.splitext(os.path.basename(file_path))[0] if file_path else "unknown"
        filename = os.path.join(save_dir, f"{filename_prefix}_{pdf_basename}_{timestamp}.png")
        image.save(filename)
        messagebox.showinfo("Download Complete", f"{filename_prefix} image saved successfully to:\n{save_dir}")
    except Exception as e:
        messagebox.showerror("Download Error", f"Failed to save {filename_prefix} image: {e}")

def download_mnist_images(debug_data, parent_window, file_path=None):
    """Prompt user to save MNIST-formatted images with unique filenames."""
    if 'mnist_images' not in debug_data or not debug_data['mnist_images']:
        messagebox.showinfo("Download Error", "No MNIST-formatted images available to download.")
        return
    
    save_dir = filedialog.askdirectory(title="Select Folder to Save MNIST Images", parent=parent_window)
    if not save_dir:
        return
    
    try:
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        pdf_basename = os.path.splitext(os.path.basename(file_path))[0] if file_path else "unknown"
        for i, mnist_img in enumerate(debug_data['mnist_images']):
            mnist_img_pil = Image.fromarray(mnist_img)
            filename = os.path.join(save_dir, f"mnist_char_{pdf_basename}_{timestamp}_{i+1}.png")
            mnist_img_pil.save(filename)
        messagebox.showinfo("Download Complete", f"MNIST images saved successfully to:\n{save_dir}")
    except Exception as e:
        messagebox.showerror("Download Error", f"Failed to save MNIST images: {e}")

# Replace the existing `show_debug_window` function with this updated version
def show_debug_window(file_path=None):
    """Show a debug window with processing details including binary classification and additional download buttons."""
    if not file_path or file_path not in debug_images:
        messagebox.showinfo("Debug Info", "No debug information available.")
        return
        
    debug_window = tk.Toplevel(root)
    debug_window.title(f"Debug Details - {os.path.basename(file_path)}")
    debug_window.geometry("800x600")
    
    main_frame = tk.Frame(debug_window)
    main_frame.pack(fill=tk.BOTH, expand=True)
    
    canvas = tk.Canvas(main_frame)
    scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
    
    scrollable_frame = tk.Frame(canvas)
    scrollable_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
    )
    
    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)
    
    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")
    
    debug_data = debug_images[file_path]
    row = 0
    
    if 'results' in debug_data:
        tk.Label(
            scrollable_frame, 
            text=f"Recognition Result: {debug_data['results']}", 
            font=("Arial", 14, "bold")
        ).grid(row=row, column=0, columnspan=2, pady=10, sticky="w", padx=10)
        row += 1
    
    # Cropped Image Section
    if 'cropped' in debug_data and debug_data['cropped']:
        cropped_frame = tk.Frame(scrollable_frame)
        cropped_frame.grid(row=row, column=0, sticky="w", padx=10, pady=5)
        
        tk.Label(
            cropped_frame, 
            text="Cropped Image:", 
            font=("Arial", 12)
        ).pack(side=tk.LEFT, padx=5)
        
        download_cropped_btn = tk.Button(
            cropped_frame, 
            text="Download", 
            command=lambda: download_image(debug_data['cropped'], "cropped", debug_window, file_path),
            width=10,
            height=1,
            bg="#ffffff",
            fg="black",
            font=("Arial", 10, "bold")
        )
        download_cropped_btn.pack(side=tk.LEFT, padx=5)
        
        row += 1
        
        img = debug_data['cropped'].copy()
        img.thumbnail((400, 400))
        tk_img = ImageTk.PhotoImage(img)
        label = tk.Label(scrollable_frame, image=tk_img)
        label.image = tk_img
        label.grid(row=row, column=0, padx=10, pady=5, sticky="w")
        row += 1
    
    # Processed Image Section
    if 'processed' in debug_data and debug_data['processed']:
        processed_frame = tk.Frame(scrollable_frame)
        processed_frame.grid(row=row, column=0, sticky="w", padx=10, pady=5)
        
        tk.Label(
            processed_frame, 
            text="Processed Image (After Preprocessing):", 
            font=("Arial", 12)
        ).pack(side=tk.LEFT, padx=5)
        
        download_processed_btn = tk.Button(
            processed_frame, 
            text="Download", 
            command=lambda: download_image(debug_data['processed'], "processed", debug_window, file_path),
            width=10,
            height=1,
            bg="#ffffff",
            fg="black",
            font=("Arial", 10, "bold")
        )
        download_processed_btn.pack(side=tk.LEFT, padx=5)
        
        row += 1
        
        img = debug_data['processed'].copy()
        img.thumbnail((400, 400))
        tk_img = ImageTk.PhotoImage(img)
        label = tk.Label(scrollable_frame, image=tk_img)
        label.image = tk_img
        label.grid(row=row, column=0, padx=10, pady=5, sticky="w")
        row += 1
    
    # Digit Detection Section with Edit button
    if 'green_boxes_img' in debug_data and debug_data['green_boxes_img']:
        detection_frame = tk.Frame(scrollable_frame)
        detection_frame.grid(row=row, column=0, sticky="w", padx=10, pady=5)
        
        tk.Label(
            detection_frame, 
            text="Handwritten Characters Detection (Green Bounding Boxes):", 
            font=("Arial", 12)
        ).pack(side=tk.LEFT, padx=5)
        
        download_detection_btn = tk.Button(
            detection_frame, 
            text="Download", 
            command=lambda: download_image(debug_data['green_boxes_img'], "digit_detection", debug_window, file_path),
            width=10,
            height=1,
            bg="#ffffff",
            fg="black",
            font=("Arial", 10, "bold")
        )
        download_detection_btn.pack(side=tk.LEFT, padx=5)
        
        # Add Edit button
        edit_btn = tk.Button(
            detection_frame, 
            text="Edit Boxes", 
            command=lambda: edit_bounding_boxes(debug_data, file_path, debug_window),
            width=10,
            height=1,
            bg="#FF9800",
            fg="black",
            font=("Arial", 10, "bold")
        )
        edit_btn.pack(side=tk.LEFT, padx=5)
        
        row += 1
        
        img = debug_data['green_boxes_img'].copy()
        img.thumbnail((400, 400))
        tk_img = ImageTk.PhotoImage(img)
        label = tk.Label(scrollable_frame, image=tk_img)
        label.image = tk_img
        label.grid(row=row, column=0, padx=10, pady=5, sticky="w")
        row += 1
    
    # MNIST Images Section
    if 'mnist_images' in debug_data and debug_data['mnist_images']:
        mnist_title_frame = tk.Frame(scrollable_frame)
        mnist_title_frame.grid(row=row, column=0, sticky="w", padx=10, pady=5)
        
        tk.Label(
            mnist_title_frame, 
            text="MNIST-Formatted Images (Used for Classification):", 
            font=("Arial", 12)
        ).pack(side=tk.LEFT, padx=5)
        
        download_btn = tk.Button(
            mnist_title_frame, 
            text="Download", 
            command=lambda: download_mnist_images(debug_data, debug_window, file_path),
            width=10,
            height=1,
            bg="#ffffff",
            fg="black",
            font=("Arial", 10, "bold")
        )
        download_btn.pack(side=tk.LEFT, padx=5)
        
        row += 1
        
        mnist_frame = tk.Frame(scrollable_frame)
        mnist_frame.grid(row=row, column=0, padx=10, pady=5, sticky="w")
        
        for i, mnist_img in enumerate(debug_data['mnist_images']):
            mnist_img_pil = Image.fromarray(mnist_img)
            mnist_img_pil = mnist_img_pil.resize((80, 80), Image.NEAREST)
            tk_img = ImageTk.PhotoImage(mnist_img_pil)
            
            digit_frame = tk.Frame(mnist_frame)
            digit_frame.pack(side=tk.LEFT, padx=5, pady=5)
            
            label = tk.Label(digit_frame, image=tk_img)
            label.image = tk_img
            label.pack()
            
            if 'binary_results' in debug_data and i < len(debug_data['binary_results']):
                tk.Label(
                    digit_frame, 
                    text=f"Type: {debug_data['binary_results'][i]}", 
                    font=("Arial", 10, "bold"),
                    fg="blue"
                ).pack()
            if 'results' in debug_data and i < len(debug_data['results']):
                tk.Label(
                    digit_frame, 
                    text=f"Result: {debug_data['results'][i]}", 
                    font=("Arial", 10, "bold")
                ).pack()
        
        row += 1

    tk.Button(scrollable_frame, text="Close", command=debug_window.destroy, width=20, height=2).grid(row=row, column=0, pady=20)

def edit_bounding_boxes(debug_data, file_path, parent_window):
    """Open a new window for editing bounding boxes on the detection image."""
    edit_window = tk.Toplevel(parent_window)
    edit_window.title(f"Edit Bounding Boxes - {os.path.basename(file_path)}")
    edit_window.geometry("1000x700")
    
    # Frame for control buttons
    control_frame = tk.Frame(edit_window)
    control_frame.pack(fill=tk.X, padx=10, pady=5)
    
    # Create a custom frame with canvas for the image editing
    image_frame = tk.Frame(edit_window, bd=2, relief=tk.SUNKEN)
    image_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
    
    # Canvas for image display with scrollbars
    h_scrollbar = ttk.Scrollbar(image_frame, orient="horizontal")
    v_scrollbar = ttk.Scrollbar(image_frame, orient="vertical")
    
    edit_canvas = tk.Canvas(
        image_frame, 
        bg="light gray",
        xscrollcommand=h_scrollbar.set,
        yscrollcommand=v_scrollbar.set
    )
    
    h_scrollbar.config(command=edit_canvas.xview)
    v_scrollbar.config(command=edit_canvas.yview)
    
    h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
    v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    edit_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    
    # Status frame for messages
    status_frame = tk.Frame(edit_window)
    status_frame.pack(fill=tk.X, padx=10, pady=5)
    
    status_label = tk.Label(
        status_frame, 
        text="Click 'Add Box' to create a new bounding box or select an existing box to remove it", 
        bd=1, 
        relief=tk.SUNKEN, 
        anchor=tk.W
    )
    status_label.pack(fill=tk.X)
    
    # Frame for buttons
    button_frame = tk.Frame(edit_window)
    button_frame.pack(fill=tk.X, padx=10, pady=5)
    
    # Variables for tracking edit state
    edit_mode = tk.StringVar(value="select")  # 'select', 'add', 'remove'
    selected_box = None
    start_x = start_y = 0
    rect_id = None
    boxes = []  # Will store [x1, y1, x2, y2, rect_id] for each box
    
    # Load initial image
    original_img = None
    try:
        # Get the processed binary image as a base to work with
        if 'processed' in debug_data and debug_data['processed']:
            # Convert PIL image to CV2 for processing
            processed_img = debug_data['processed']
            np_img = np.array(processed_img)
            if len(np_img.shape) == 2:  # Grayscale
                original_img = cv2.cvtColor(np_img, cv2.COLOR_GRAY2BGR)
            else:
                original_img = cv2.cvtColor(np_img, cv2.COLOR_RGB2BGR)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to load image for editing: {e}")
        edit_window.destroy()
        return
    
    # Get existing bounding boxes if available
    existing_boxes = []
    if file_path in debug_images and 'bounding_boxes' in debug_images[file_path]:
        existing_boxes = debug_images[file_path]['bounding_boxes']
    
    # Function to draw all boxes on the image
    def redraw_image():
        nonlocal original_img, boxes
        
        # Create a clean copy of the original image
        display_img = original_img.copy()
        
        # Draw all the saved boxes
        for x1, y1, x2, y2, _ in boxes:
            cv2.rectangle(display_img, (x1, y1), (x2, y2), (0, 255, 0), 2)
        
        # Convert back to PIL for display
        img_rgb = cv2.cvtColor(display_img, cv2.COLOR_BGR2RGB)
        pil_img = Image.fromarray(img_rgb)
        
        # Update the canvas
        width, height = pil_img.size
        edit_canvas.config(scrollregion=(0, 0, width, height))
        
        # Store the image references to prevent garbage collection
        edit_window.pil_img = pil_img
        edit_window.tk_img = ImageTk.PhotoImage(pil_img)
        
        # Display the image
        edit_canvas.delete("all")
        edit_canvas.create_image(0, 0, anchor="nw", image=edit_window.tk_img)
        
        # Redraw the boxes as selectable rectangles
        for i, (x1, y1, x2, y2, _) in enumerate(boxes):
            rect_id = edit_canvas.create_rectangle(
                x1, y1, x2, y2, 
                outline="red", 
                width=2,
                tags=(f"box_{i}", "box")
            )
            # Update the rect_id in the boxes list
            boxes[i] = [x1, y1, x2, y2, rect_id]
        
        # Bind box selection
        edit_canvas.tag_bind("box", "<ButtonPress-1>", on_select_box)
    
    # Initialize with existing boxes
    for x, y, w, h in existing_boxes:
        boxes.append([x, y, x+w, y+h, None])
    
    # Event handlers for box editing
    def on_select_box(event):
        nonlocal selected_box
        if edit_mode.get() != "select":
            return
            
        # Find which box was clicked
        clicked_item = edit_canvas.find_closest(edit_canvas.canvasx(event.x), edit_canvas.canvasy(event.y))[0]
        tags = edit_canvas.gettags(clicked_item)
        
        # Check if it's a box
        box_tag = next((tag for tag in tags if tag.startswith("box_")), None)
        if box_tag:
            # Extract box index
            box_idx = int(box_tag.split("_")[1])
            selected_box = box_idx
            
            # Highlight selected box
            for i, (x1, y1, x2, y2, rect_id) in enumerate(boxes):
                if i == box_idx:
                    edit_canvas.itemconfig(rect_id, outline="blue", width=3)
                else:
                    edit_canvas.itemconfig(rect_id, outline="red", width=2)
            
            status_label.config(text=f"Box {box_idx} selected. Click 'Remove Box' to delete it.")
    
    def start_add_box(event):
        nonlocal start_x, start_y, rect_id
        if edit_mode.get() != "add":
            return
            
        # Get canvas coordinates
        canvas_x = edit_canvas.canvasx(event.x)
        canvas_y = edit_canvas.canvasy(event.y)
        
        start_x, start_y = int(canvas_x), int(canvas_y)
        rect_id = edit_canvas.create_rectangle(
            canvas_x, canvas_y, canvas_x, canvas_y, 
            outline="green", 
            width=2
        )
    
    def update_add_box(event):
        nonlocal rect_id
        if edit_mode.get() != "add" or rect_id is None:
            return
            
        # Get canvas coordinates
        canvas_x = edit_canvas.canvasx(event.x)
        canvas_y = edit_canvas.canvasy(event.y)
        
        # Update rectangle
        edit_canvas.coords(rect_id, start_x, start_y, canvas_x, canvas_y)
    
    def finish_add_box(event):
        nonlocal rect_id, boxes
        if edit_mode.get() != "add" or rect_id is None:
            return
            
        # Get canvas coordinates
        canvas_x = edit_canvas.canvasx(event.x)
        canvas_y = edit_canvas.canvasy(event.y)
        
        # Ensure we have a valid box (minimum size)
        end_x, end_y = int(canvas_x), int(canvas_y)
        x1, y1 = min(start_x, end_x), min(start_y, end_y)
        x2, y2 = max(start_x, end_x), max(start_y, end_y)
        
        if (x2 - x1) > 5 and (y2 - y1) > 5:  # Minimum size check
            # Add to boxes list
            boxes.append([x1, y1, x2, y2, rect_id])
            status_label.config(text=f"Box added at ({x1}, {y1}, {x2}, {y2})")
        else:
            # Too small, delete it
            edit_canvas.delete(rect_id)
            status_label.config(text="Box too small. Please try again.")
        
        rect_id = None
        set_mode("select")
    
    def set_mode(mode):
        """Set the current edit mode and update UI accordingly"""
        edit_mode.set(mode)
        
        # Remove any bindings
        edit_canvas.unbind("<ButtonPress-1>")
        edit_canvas.unbind("<B1-Motion>")
        edit_canvas.unbind("<ButtonRelease-1>")
        
        # Unbind box selection
        edit_canvas.tag_unbind("box", "<ButtonPress-1>")
        
        # Set cursor and bindings based on mode
        if mode == "add":
            edit_canvas.config(cursor="cross")
            status_label.config(text="Click and drag to create a new bounding box")
            edit_canvas.bind("<ButtonPress-1>", start_add_box)
            edit_canvas.bind("<B1-Motion>", update_add_box)
            edit_canvas.bind("<ButtonRelease-1>", finish_add_box)
            
            # Update button states
            add_box_btn.config(bg="#FF9800", state=tk.DISABLED)
            remove_box_btn.config(bg="#ffffff", state=tk.NORMAL)
            select_mode_btn.config(bg="#ffffff", state=tk.NORMAL)
        elif mode == "select":
            edit_canvas.config(cursor="arrow")
            status_label.config(text="Click on a box to select it")
            # Bind box selection
            edit_canvas.tag_bind("box", "<ButtonPress-1>", on_select_box)
            
            # Update button states
            add_box_btn.config(bg="#ffffff", state=tk.NORMAL)
            remove_box_btn.config(bg="#ffffff", state=tk.NORMAL)
            select_mode_btn.config(bg="#FF9800", state=tk.DISABLED)
    
    def remove_selected_box():
        nonlocal selected_box, boxes
        if selected_box is not None and 0 <= selected_box < len(boxes):
            # Remove the box from canvas
            _, _, _, _, rect_id = boxes[selected_box]
            edit_canvas.delete(rect_id)
            
            # Remove from boxes list
            boxes.pop(selected_box)
            status_label.config(text=f"Box {selected_box} removed")
            
            # Reset selection
            selected_box = None
            
            # Redraw to update box indices
            redraw_image()
        else:
            status_label.config(text="No box selected")
    
    def save_and_close():
        # Convert boxes to OpenCV format (x, y, w, h)
        cv_boxes = []
        for x1, y1, x2, y2, _ in sorted(boxes, key=lambda b: b[0]):  # Sort by x-coordinate
            x = min(x1, x2)
            y = min(y1, y2)
            w = abs(x2 - x1)
            h = abs(y2 - y1)
            cv_boxes.append((x, y, w, h))
        
        # Save boxes back to debug_images
        debug_images[file_path]['bounding_boxes'] = cv_boxes
        
        # Update the green_boxes_img in debug_images
        display_img = original_img.copy()
        for x, y, w, h in cv_boxes:
            cv2.rectangle(display_img, (x, y), (x+w, y+h), (0, 255, 0), 2)
        
        img_rgb = cv2.cvtColor(display_img, cv2.COLOR_BGR2RGB)
        pil_img = Image.fromarray(img_rgb)
        debug_images[file_path]['green_boxes_img'] = pil_img
        
        # Process the edited boxes to update results
        update_recognition_results(debug_data, file_path, cv_boxes)

        # Update recognition results for the results window
        recognition_results = []
        for fp in pdf_files:
            if fp in debug_images and 'results' in debug_images[fp]:
                recognition_results.append(f"{os.path.basename(fp)}: {debug_images[fp]['results']}")
            else:
                recognition_results.append(f"{os.path.basename(fp)}: Not processed")
        
        # Refresh the results window if it exists
        if result_window is not None and result_window.winfo_exists():
            # Update the text area
            result_window.text_area.configure(state="normal")
            result_window.text_area.delete(1.0, tk.END)
            for result in recognition_results:
                result_window.text_area.insert(tk.END, result + "\n")
            result_window.text_area.configure(state="disabled")
        else:
            # If no results window, create a new one
            show_recognition_results(recognition_results)
        
        messagebox.showinfo("Success", "Bounding boxes updated successfully")
        edit_window.destroy()
        
        # Refresh the debug window to show updated boxes
        parent_window.destroy()
        show_debug_window(file_path)
    
    # Button for selecting mode
    select_mode_btn = tk.Button(
        control_frame,
        text="Select Mode",
        command=lambda: set_mode("select"),
        width=12,
        height=1,
        bg="#FF9800",  # Initially selected
        fg="black",
        font=("Arial", 10, "bold")
    )
    select_mode_btn.pack(side=tk.LEFT, padx=5)
    
    # Button for adding a new box
    add_box_btn = tk.Button(
        control_frame,
        text="Add Box",
        command=lambda: set_mode("add"),
        width=12,
        height=1,
        bg="#ffffff",
        fg="black",
        font=("Arial", 10, "bold")
    )
    add_box_btn.pack(side=tk.LEFT, padx=5)
    
    # Button for removing a selected box
    remove_box_btn = tk.Button(
        control_frame,
        text="Remove Box",
        command=remove_selected_box,
        width=12,
        height=1,
        bg="#ffffff",
        fg="black",
        font=("Arial", 10, "bold")
    )
    remove_box_btn.pack(side=tk.LEFT, padx=5)
    
    # Save and cancel buttons
    save_btn = tk.Button(
        button_frame,
        text="Save Changes",
        command=save_and_close,
        width=15,
        height=2,
        bg="#4CAF50",
        fg="black",
        font=("Arial", 10, "bold")
    )
    save_btn.pack(side=tk.RIGHT, padx=5)
    
    cancel_btn = tk.Button(
        button_frame,
        text="Cancel",
        command=edit_window.destroy,
        width=15,
        height=2,
        bg="#f44336",
        fg="black",
        font=("Arial", 10, "bold")
    )
    cancel_btn.pack(side=tk.RIGHT, padx=5)
    
    # Draw initial image with boxes
    redraw_image()
    
    # Set initial mode
    set_mode("select")

def update_recognition_results(debug_data, file_path, bounding_boxes):
    """Process the edited bounding boxes to update recognition results."""
    try:
        # Get the processed image
        processed_img = debug_data['processed']
        np_img = np.array(processed_img)
        
        # Create a temporary directory for MNIST-formatted images
        temp_dir = "/Users/ryan/Desktop/temp_edited_digits"
        os.makedirs(temp_dir, exist_ok=True)
        
        # Sort bounding boxes by x-coordinate for left-to-right order
        sorted_boxes = sorted(bounding_boxes, key=lambda box: box[0])
        
        # Extract, process and classify each character from sorted bounding boxes
        digit_results = []
        mnist_images = []
        binary_results = []
        
        for i, (x, y, w, h) in enumerate(sorted_boxes):
            # Extract the character region
            char_roi = np_img[y:y+h, x:x+w]
            
            # Convert to MNIST format
            mnist_digit = preprocess_to_mnist_format(char_roi)
            mnist_images.append(mnist_digit)
            
            # Save for classification
            digit_filename = os.path.join(temp_dir, f"edited_char_{i+1}.png")
            cv2.imwrite(digit_filename, mnist_digit)
            
            # Classify the character
            result, binary_label = classify_image(digit_filename)
            if result is not None:
                digit_results.append(result)
                binary_results.append(binary_label)
            else:
                digit_results.append("?")
                binary_results.append("Unknown")
        
        # Update debug data with new results
        debug_data.update({
            'mnist_images': mnist_images,
            'results': ''.join(digit_results),
            'binary_results': binary_results,
            'bounding_boxes': sorted_boxes
        })
        
        # Clean up temporary directory
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
            
    except Exception as e:
        messagebox.showerror("Processing Error", f"Error updating recognition results: {e}")
        # Update with empty results to prevent further errors
        debug_data.update({
            'mnist_images': [],
            'results': '',
            'binary_results': [],
            'bounding_boxes': sorted_boxes
        })

# Helper function for processing MNIST format images
def preprocess_to_mnist_format(image):
    """Convert an image to MNIST format (28x28, centered, grayscale)."""
    if len(image.shape) > 2:
        image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    
    # Ensure the image is binary
    _, image = cv2.threshold(image, 128, 255, cv2.THRESH_BINARY)
    
    # Find non-zero (foreground) coordinates
    coords = cv2.findNonZero(image)
    if coords is None:
        return np.zeros((28, 28), dtype=np.uint8)
    
    # Get bounding rect of character
    x, y, w, h = cv2.boundingRect(coords)
    if w == 0 or h == 0:
        return np.zeros((28, 28), dtype=np.uint8)
    
    # Extract the character region
    digit = image[y:y+h, x:x+w]
    
    # Calculate the scale to fit within 20x20 area
    max_dim = max(w, h)
    scale = 20.0 / max_dim
    new_w = int(w * scale)
    new_h = int(h * scale)
    
    # Resize character to fit within 20x20 area
    resized = cv2.resize(digit, (new_w, new_h), interpolation=cv2.INTER_LINEAR)
    
    # Create 28x28 canvas and center character
    canvas = np.zeros((28, 28), dtype=np.uint8)
    x_offset = (28 - new_w) // 2
    y_offset = (28 - new_h) // 2
    canvas[y_offset:y_offset+new_h, x_offset:x_offset+new_w] = resized
    
    return canvas

def setup_canvas():
    """Configure the canvas when the window is resized."""
    if pdf_image:
        display_image(pdf_image)

def on_combobox_select(event):
    """Handle selection change in the PDF combobox."""
    display_selected_pdf()

# Initialize the main application
root = tk.Tk()
root.title("New HRC - Handwritten Recognition")
root.state('zoomed')

try:
    style = ttk.Style()
    available_themes = style.theme_names()
    if 'clam' in available_themes:
        style.theme_use('clam')
    elif 'vista' in available_themes:
        style.theme_use('vista')
except Exception:
    pass

# Global variables (removed replace_var and user_entry_var)
pdf_image = None
tk_image = None
scaled_image = None
image_scale = 1
start_x = start_y = 0
rect_id = None
crop_coords = None
pdf_files = []
debug_images = {}
canvas_offset_x = canvas_offset_y = 0
zoom_factor = 1.0
original_image = None
zoom_mode = False
pan_start_x = None
pan_start_y = None
result_window = None  # Add this to the global variables section

main_frame = tk.Frame(root)
main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

control_frame = tk.Frame(main_frame)
control_frame.pack(fill=tk.X, pady=10)

upload_button = tk.Button(control_frame, text="Upload PDFs", command=load_pdf, width=15, height=2)
upload_button.grid(row=0, column=0, padx=5)

pdf_label = tk.Label(control_frame, text="Select PDF:")
pdf_label.grid(row=0, column=1, padx=5, pady=5, sticky="e")

pdf_combobox = ttk.Combobox(control_frame, width=40, state="readonly")
pdf_combobox.grid(row=0, column=2, padx=5, pady=5, sticky="w")
pdf_combobox.bind("<<ComboboxSelected>>", on_combobox_select)

classify_button = tk.Button(
    control_frame, 
    text="Classify", 
    command=classify,
    width=15,
    height=2,
    bg="#4CAF50",
    fg="black",
    font=("Arial", 10, "bold")
)
classify_button.grid(row=0, column=3, padx=5, pady=5)

status_label = tk.Label(
    control_frame, 
    text="Ready. Please load a PDF file.", 
    font=("Arial", 10),
    bd=1,
    relief=tk.SUNKEN,
    anchor=tk.W
)
status_label.grid(row=0, column=4, columnspan=1, padx=5, pady=5, sticky="ew")

canvas_frame = tk.Frame(main_frame, bd=2, relief=tk.SUNKEN)
canvas_frame.pack(fill=tk.BOTH, expand=True, pady=10)

canvas = tk.Canvas(canvas_frame, bg="light gray")
canvas.pack(fill=tk.BOTH, expand=True)

instructions = tk.Label(
    main_frame, 
    text="Instructions: Upload a PDF, select an area with digits by drawing a rectangle, then click 'Classify'",
    font=("Arial", 10),
    fg="#666666"
)
instructions.pack(pady=5, anchor=tk.W)

# In your control_frame setup, add zoom controls
zoom_button = tk.Button(
    control_frame, 
    text="Switch to Zoom Mode", 
    command=toggle_zoom_mode,
    width=15,
    height=2,
    bg="#2196F3",
    fg="black",
    font=("Arial", 10, "bold")
)
zoom_button.grid(row=0, column=5, padx=5, pady=5)

reset_zoom_button = tk.Button(
    control_frame, 
    text="Reset Zoom", 
    command=reset_zoom,
    width=10,
    height=2,
    bg="#9E9E9E",
    fg="black",
    font=("Arial", 10, "bold")
)
reset_zoom_button.grid(row=0, column=6, padx=5, pady=5)

# Configure the canvas for scrolling when zoomed
h_scrollbar = ttk.Scrollbar(canvas_frame, orient="horizontal", command=canvas.xview)
v_scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=canvas.yview)
canvas.configure(xscrollcommand=h_scrollbar.set, yscrollcommand=v_scrollbar.set)

# Pack all elements in the canvas_frame
canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)

canvas.bind("<ButtonPress-1>", start_crop)
canvas.bind("<B1-Motion>", update_crop)
canvas.bind("<ButtonRelease-1>", finish_crop)
canvas.bind("<Configure>", lambda e: setup_canvas())

root.mainloop()
