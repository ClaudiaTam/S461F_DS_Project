import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk
from PIL import Image, ImageTk
from pdf2image import convert_from_path
import os
import logging
import tempfile
from image_processing import preprocess_image, detect_characters
from classification import classify_characters, model
from utils import pdf_to_image
import openpyxl

# Setup logging
logging.basicConfig(
    filename='handwritten_recognition.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

class HandwrittenRecognitionApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Deep Learning for Handwritten Characters Recognition")
        self.root.state('zoomed')
        self.pdf_image = None
        self.tk_image = None
        self.scaled_image = None
        self.image_scale = 1
        self.start_x = self.start_y = 0
        self.rect_id = None
        self.crop_coords_dict = {}  # Store crop coords and labels per PDF
        self.pdf_files = []
        self.debug_images = {}
        self.setup_gui()

    def setup_gui(self):
        main_frame = tk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        control_frame = tk.Frame(main_frame)
        control_frame.pack(fill=tk.X, pady=10)

        tk.Button(control_frame, text="Upload PDFs", command=self.load_pdf, width=15, height=2).grid(row=0, column=0, padx=5)
        tk.Label(control_frame, text="Select PDF:").grid(row=0, column=1, padx=5, pady=5, sticky="e")
        self.pdf_combobox = ttk.Combobox(control_frame, width=40, state="readonly")
        self.pdf_combobox.grid(row=0, column=2, padx=5, pady=5, sticky="w")
        self.pdf_combobox.bind("<<ComboboxSelected>>", self.on_combobox_select)

        tk.Label(control_frame, text="Label Crop Area:").grid(row=1, column=0, padx=5, pady=5, sticky="e")
        self.label_var = tk.StringVar(value="course_code")
        tk.Radiobutton(control_frame, text="Course Code", variable=self.label_var, value="course_code").grid(row=1, column=1, padx=5, pady=5, sticky="w")
        tk.Radiobutton(control_frame, text="Exam/Student Number", variable=self.label_var, value="exam_student_number").grid(row=1, column=2, padx=5, pady=5, sticky="w")

        tk.Button(control_frame, text="Classify", command=self.classify, width=15, height=2, bg="#4CAF50", fg="black", font=("Arial", 10, "bold")).grid(row=2, column=0, padx=5, pady=5)

        self.status_label = tk.Label(control_frame, text="Ready. Please load a PDF file.", font=("Arial", 10), bd=1, relief=tk.SUNKEN, anchor=tk.W)
        self.status_label.grid(row=2, column=1, columnspan=2, padx=5, pady=5, sticky="ew")

        tk.Button(control_frame, text="Help", command=self.show_help).grid(row=0, column=3, padx=5)

        canvas_frame = tk.Frame(main_frame, bd=2, relief=tk.SUNKEN)
        canvas_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        self.canvas = tk.Canvas(canvas_frame, bg="light gray")
        self.canvas.pack(fill=tk.BOTH, expand=True)

        tk.Label(main_frame, text="Instructions: Upload a PDF, select an area, label it, and click 'Classify'", font=("Arial", 10), fg="#666666").pack(pady=5, anchor=tk.W)

        self.canvas.bind("<ButtonPress-1>", self.start_crop)
        self.canvas.bind("<B1-Motion>", self.update_crop)
        self.canvas.bind("<ButtonRelease-1>", self.finish_crop)
        self.canvas.bind("<Configure>", self.on_canvas_resize)

    def show_help(self):
        messagebox.showinfo("Help", "1. Upload PDFs using 'Upload PDFs'.\n2. Select a PDF from the dropdown.\n3. Draw a rectangle over the text to classify.\n4. Label the area as 'Course Code' or 'Exam/Student Number'.\n5. Click 'Classify' to process.")

    def load_pdf(self):
        file_paths = filedialog.askopenfilenames(filetypes=[("PDF Files", "*.pdf")], title="Select one or more one-page PDF files")
        if not file_paths:
            return
        self.pdf_files = list(file_paths)
        self.debug_images.clear()
        self.crop_coords_dict.clear()
        self.pdf_combobox['values'] = [os.path.basename(path) for path in self.pdf_files]
        if self.pdf_files:
            self.pdf_combobox.current(0)
            self.display_selected_pdf()
        logging.info(f"Loaded PDFs: {self.pdf_files}")

    def display_selected_pdf(self):
        selection = self.pdf_combobox.current()
        if selection < 0 or not self.pdf_files:
            return
        self.pdf_image = pdf_to_image(self.pdf_files[selection])
        if self.pdf_image:
            self.display_image(self.pdf_image)
            self.status_label.config(text=f"Loaded: {os.path.basename(self.pdf_files[selection])}")
            logging.info(f"Displayed PDF: {self.pdf_files[selection]}")

    def display_image(self, image):
        canvas_width, canvas_height = self.canvas.winfo_width(), self.canvas.winfo_height()
        if canvas_width <= 1:
            canvas_width, canvas_height = 600, 800
        scale_factor = min(canvas_width / image.width, canvas_height / image.height)
        new_width, new_height = int(image.width * scale_factor), int(image.height * scale_factor)
        self.scaled_image = image.resize((new_width, new_height))
        self.image_scale = scale_factor
        self.canvas_offset_x = (canvas_width - new_width) // 2
        self.canvas_offset_y = (canvas_height - new_height) // 2
        self.tk_image = ImageTk.PhotoImage(self.scaled_image)
        self.canvas.delete("all")
        self.canvas.create_image(self.canvas_offset_x, self.canvas_offset_y, anchor="nw", image=self.tk_image)

    def start_crop(self, event):
        if self.rect_id:
            self.canvas.delete(self.rect_id)
        self.start_x, self.start_y = event.x - self.canvas_offset_x, event.y - self.canvas_offset_y
        self.rect_id = self.canvas.create_rectangle(event.x, event.y, event.x, event.y, outline="red", width=2)

    def update_crop(self, event):
        end_x = max(min(event.x, self.canvas.winfo_width()), 0)
        end_y = max(min(event.y, self.canvas.winfo_height()), 0)
        self.canvas.coords(self.rect_id, self.start_x + self.canvas_offset_x, self.start_y + self.canvas_offset_y, end_x, end_y)

    def finish_crop(self, event):
        if not self.pdf_image:
            messagebox.showerror("Error", "No PDF image loaded.")
            return
        end_x, end_y = event.x - self.canvas_offset_x, event.y - self.canvas_offset_y
        x1, y1 = min(self.start_x, end_x), min(self.start_y, end_y)
        x2, y2 = max(self.start_x, end_x), max(self.start_y, end_y)
        x1 = max(0, min(x1, self.scaled_image.width)) / self.image_scale
        y1 = max(0, min(y1, self.scaled_image.height)) / self.image_scale
        x2 = max(0, min(x2, self.scaled_image.width)) / self.image_scale
        y2 = max(0, min(y2, self.scaled_image.height)) / self.image_scale
        selected_pdf = self.pdf_files[self.pdf_combobox.current()]
        self.crop_coords_dict[selected_pdf] = {
            'coords': (int(x1), int(y1), int(x2), int(y2)),
            'label': self.label_var.get()
        }
        self.status_label.config(text=f"Selection complete for {self.label_var.get()}. Ready to classify.")
        logging.info(f"Crop area set for {selected_pdf}: {self.label_var.get()}")

    def on_canvas_resize(self, event):
        if self.pdf_image:
            self.display_image(self.pdf_image)

    def process_single_pdf(self, file_path, coords):
        """Process a single PDF and return recognition result."""
        original_img = pdf_to_image(file_path)
        if not original_img:
            raise ValueError("Failed to convert PDF to image")
        cropped_image = original_img.crop(coords)
        with tempfile.NamedTemporaryFile(suffix='.png') as temp_file:
            cropped_image.save(temp_file.name)
            binary_img, cleaned_binary = preprocess_image(temp_file.name)
            bounding_boxes, mnist_images, green_boxes_img = detect_characters(binary_img, cleaned_binary)
            char_results = classify_characters(mnist_images, model)
        return char_results, {
            'original': original_img,
            'cropped': cropped_image,
            'processed': Image.fromarray(cleaned_binary),
            'mnist_images': mnist_images,
            'green_boxes_img': Image.fromarray(green_boxes_img),
            'results': ''.join(char_results)
        }

    def classify(self):
        if not self.pdf_files:
            messagebox.showerror("Error", "No PDFs loaded.")
            return
        self.status_label.config(text="Processing... Please wait.")
        self.root.update()
        if self.rect_id:
            self.canvas.delete(self.rect_id)
            self.rect_id = None

        recognition_results = []
        course_codes = {}
        exam_student_numbers = {}
        for file_path in self.pdf_files:
            try:
                entry = self.crop_coords_dict.get(file_path)
                if not entry:
                    raise ValueError("No crop area selected for this PDF")
                coords, label = entry['coords'], entry['label']
                logging.info(f"Processing {file_path} as {label}")
                char_results, debug_entry = self.process_single_pdf(file_path, coords)
                result = ''.join(char_results)
                self.debug_images[file_path] = debug_entry
                if label == "course_code":
                    course_codes[file_path] = result
                else:
                    exam_student_numbers[file_path] = result
                recognition_results.append(f"{os.path.basename(file_path)} ({label}): {result}")
                logging.info(f"Result for {file_path}: {result}")
            except Exception as e:
                error_msg = f"{os.path.basename(file_path)}: Error - {e}"
                retry = messagebox.askretrycancel("Processing Error", f"{error_msg}\nRetry?")
                if retry:
                    self.classify()
                    return
                else:
                    recognition_results.append(error_msg)
                    logging.error(error_msg)

        # Organize results for Excel export
        excel_data = []
        for file_path in self.pdf_files:
            filename = os.path.basename(file_path)
            course_code = course_codes.get(file_path, "N/A")
            exam_number = exam_student_numbers.get(file_path, "N/A")
            excel_data.append((filename, course_code, exam_number))

        self.status_label.config(text="Classification complete!")
        self.show_recognition_results(recognition_results, excel_data)

    def show_recognition_results(self, results, excel_data):
        result_window = tk.Toplevel(self.root)
        result_window.title("Recognition Results")
        result_window.geometry("1200x400")

        control_frame = tk.Frame(result_window)
        control_frame.pack(fill=tk.X, padx=10, pady=5)

        tk.Label(control_frame, text="Select PDF for Debug:").pack(side=tk.LEFT, padx=5)
        pdf_debug_combobox = ttk.Combobox(control_frame, width=40, state="readonly")
        pdf_debug_combobox.pack(side=tk.LEFT, padx=5)
        pdf_debug_combobox['values'] = [os.path.basename(path) for path in self.pdf_files]
        if self.pdf_files:
            pdf_debug_combobox.current(self.pdf_combobox.current())

        tk.Button(control_frame, text="Show Debug Details", command=lambda: self.show_debug_window(pdf_debug_combobox.get())).pack(side=tk.LEFT, padx=5)
        tk.Button(control_frame, text="Output to Excel", command=lambda: self.export_results_to_excel(excel_data, result_window)).pack(side=tk.LEFT, padx=5)
        tk.Button(control_frame, text="Close", command=result_window.destroy).pack(side=tk.RIGHT, padx=5)

        text_area = scrolledtext.ScrolledText(result_window, wrap=tk.WORD, width=60, height=20, font=("Arial", 14))
        text_area.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        for result in results:
            text_area.insert(tk.END, result + "\n")
        text_area.configure(state="disabled")

    def export_results_to_excel(self, excel_data, parent_window):
        save_dir = filedialog.askdirectory(title="Select Folder to Save Excel File", parent=parent_window)
        if not save_dir:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Recognition Results"
            ws['A1'], ws['B1'], ws['C1'] = "File Name", "Course Code", "Exam/Student Number"
            for row, (filename, course_code, exam_number) in enumerate(excel_data, start=2):
                ws[f'A{row}'] = filename
                ws[f'B{row}'] = course_code
                ws[f'C{row}'] = exam_number
            timestamp = time.strftime("%Y%m%d_%H%M%S")
            filename = os.path.join(save_dir, f"recognition_results_{timestamp}.xlsx")
            wb.save(filename)
            messagebox.showinfo("Export Complete", f"Results exported to:\n{filename}")
            logging.info(f"Exported results to {filename}")
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export: {e}")
            logging.error(f"Export error: {e}")

    def show_debug_window(self, selected_pdf):
        file_path = [p for p in self.pdf_files if os.path.basename(p) == selected_pdf][0]
        if file_path not in self.debug_images:
            messagebox.showinfo("Debug Info", "No debug information available.")
            return

        debug_window = tk.Toplevel(self.root)
        debug_window.title(f"Debug Details - {selected_pdf}")
        debug_window.geometry("800x600")

        canvas = tk.Canvas(debug_window)
        scrollbar = ttk.Scrollbar(debug_window, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas)
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        debug_data = self.debug_images[file_path]
        row = 0

        tk.Label(scrollable_frame, text=f"Recognition Result: {debug_data['results']}", font=("Arial", 14, "bold")).grid(row=row, column=0, columnspan=2, pady=10, sticky="w", padx=10)
        row += 1

        for key, title in [('cropped', 'Cropped Image'), ('processed', 'Processed Image'), ('green_boxes_img', 'Digit Detection')]:
            if key in debug_data:
                tk.Label(scrollable_frame, text=f"{title}:", font=("Arial", 12)).grid(row=row, column=0, sticky="w", padx=10, pady=5)
                img = debug_data[key].copy()
                img.thumbnail((400, 400))
                tk_img = ImageTk.PhotoImage(img)
                label = tk.Label(scrollable_frame, image=tk_img)
                label.image = tk_img
                label.grid(row=row+1, column=0, padx=10, pady=5, sticky="w")
                row += 2

        if 'mnist_images' in debug_data:
            tk.Label(scrollable_frame, text="MNIST-Formatted Images:", font=("Arial", 12)).grid(row=row, column=0, sticky="w", padx=10, pady=5)
            row += 1
            mnist_frame = tk.Frame(scrollable_frame)
            mnist_frame.grid(row=row, column=0, padx=10, pady=5, sticky="w")
            for i, mnist_img in enumerate(debug_data['mnist_images']):
                img = Image.fromarray(mnist_img).resize((80, 80), Image.NEAREST)
                tk_img = ImageTk.PhotoImage(img)
                digit_frame = tk.Frame(mnist_frame)
                digit_frame.pack(side=tk.LEFT, padx=5, pady=5)
                tk.Label(digit_frame, image=tk_img).pack()
                tk.Label(digit_frame, text=f"Result: {debug_data['results'][i]}", font=("Arial", 10, "bold")).pack()
                label.image = tk_img
            row += 1

        tk.Button(scrollable_frame, text="Close", command=debug_window.destroy, width=20, height=2).grid(row=row, column=0, pady=20)

    def on_combobox_select(self, event):
        self.display_selected_pdf()

if __name__ == "__main__":
    root = tk.Tk()
    app = HandwrittenRecognitionApp(root)
    root.mainloop()